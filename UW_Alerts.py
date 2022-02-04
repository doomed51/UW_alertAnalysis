from datetime import datetime, date
from importlib.resources import path
from os import P_DETACH, replace
from turtle import color, left
from matplotlib.colors import LinearSegmentedColormap
from numpy.core.defchararray import index
from numpy.core.fromnumeric import partition
from numpy.lib.npyio import load
from numpy.testing._private.nosetester import run_module_suite
from openpyxl import workbook, load_workbook
from openpyxl.worksheet import worksheet
from sklearn.cluster import KMeans
from pathlib import Path

import pandas as pd 
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import time
#pd.options.mode.chained_assignment = 'raise'

#####
# Read in alerts excel as dataframe 
# returns a clean dataframe 
#####
def getAlerts(symbolOrSheetName = 'all symbols'):
    # Location of the file that contains the alerts dump
    filePaths = [r"F:\workbench\UW_Alerts\UW_alerts_symbols.xlsx", 
        r"F:\workbench\UW_Alerts\UW_alerts.xlsx",
        r"F:\workbench\UW_Alerts\UW_alerts_myHunt.xlsx"]
    print('')
    print('#########################################################')
    print('Loading files for ...%s'%(symbolOrSheetName))
    start = time.perf_counter()
    if symbolOrSheetName == 'all symbols':
        print('loading worksheet...%s'%(filePaths[0]))
        workbook_symbol = load_workbook(filename=filePaths[0])
        alertsDF_map = pd.read_excel(filePaths[0], sheet_name=None)
    
        print('')
        print('Success!')
        end = time.perf_counter()
        print(f"Elapsed Time: {end - start:0.4f} seconds")
        print('#########################################################')
        print('')
        return alertsDF_map

    elif symbolOrSheetName == 'All Alerts':
        print('loading worksheet...%s'%(filePaths[1]))
        workbook = load_workbook(filename=filePaths[1])
        alertsDF_map = pd.read_excel(filePaths[1], sheet_name=None)

        print('')
        print('Success!')
        end = time.perf_counter()
        print(f"Elapsed Time: {end - start:0.4f} seconds")
        print('#########################################################')
        print('')
        return cleanAlertsData(alertsDF_map[symbolOrSheetName])
    
    elif symbolOrSheetName == 'My Hunt':
        print('loading worksheet...%s'%(filePaths[2]))
        workbook = load_workbook(filename=filePaths[2])
        alertsDF_map = pd.read_excel(filePaths[2], sheet_name=None)

        print('')
        print('Success!')
        end = time.perf_counter()
        print(f"Elapsed Time: {end - start:0.4f} seconds")
        print('#########################################################')
        print('')
        return cleanAlertsData(alertsDF_map[symbolOrSheetName])

    else:
        print('loading worksheet...ALL!!')
        workbook_symbol = load_workbook(filename=filePaths[0])
        workbook = load_workbook(filename=filePaths[1])
        workbook_myHunt = load_workbook(filename=filePaths[2])

        if symbolOrSheetName in workbook_symbol.sheetnames:
            print('Sheet found, loading alerts ...')
            alertsDF_map = pd.read_excel(filePaths[0], sheet_name=None)
        elif symbolOrSheetName in workbook.sheetnames:
            print('Sheet found, loading alerts ...')
            alertsDF_map = pd.read_excel(filePaths[1], sheet_name=None)
        elif symbolOrSheetName in workbook_myHunt.sheetnames:
            print('Sheet found, loading alerts ...')
            alertsDF_map = pd.read_excel(filePaths[2], sheet_name=None)
        else:
            print("Sheet not found!!")
            exit()

        print('')
        print('Success!')
        end = time.perf_counter()
        print(f"Elapsed Time: {end - start:0.4f} seconds")
        print('#########################################################')
        print('')

        return cleanAlertsData(alertsDF_map[symbolOrSheetName])

#####
# Objective:    Print passed in dataframe to passed in filename
#####
def printAlertsToFile(alertsDF, filename='genericAlert.csv'):
    print('')
    
    filepath = Path('output/'+filename+'.csv')
    print('Printing file: %s'%(filepath))
    filepath.parent.mkdir(parents=True, exist_ok=True)
    alertsDF.to_csv(filepath, index=False)
    
#####
# Util function
##
# replace 'C' and 'P' with 'Call' and 'Put'
#####
def replaceText(row):
    strike = row['Strike'][-1]
    if strike == 'C':
        return 'Call'
    else:
        return 'Put'

#####
# Util function
##
# Remove the trailing char in a string 
#####
def cleanStrike(row):
    return row['Strike'][:-1]

#####
# Cleans up the alerts extract - NaNs etc. 
# splits up option chain column into symbol, strike and option type
# Splits up Max gain and loss columns into % and abs. 
#####
def cleanAlertsData(alertsDF):
    # drop: Action, Emojis columns
    alertsDF = alertsDF.drop(columns=['Actions'], axis=1) 

    # clean up Time column i.e. alert time, leaving only the date
    
    alertsDF['Alert Date'] = alertsDF['Time'].astype(str).str[:-9]
    alertsDF['Alert Date'] = pd.to_datetime(alertsDF['Alert Date'], dayfirst=True)
    alertsDF['Expiry'] = pd.to_datetime(alertsDF['Expiry'], dayfirst=True)

    # remove trailing spaces & recast 
    alertsDF['Option'].str.strip()
    alertsDF['Option'] = alertsDF['Option'].astype('str')
    alertsDF['Option'] = alertsDF['Option'].str.strip()

    #split the Option col into symbol and strike
    alertsDF[['Option', 'Strike']] = alertsDF['Option'].str.split('$', expand=True)
    alertsDF['Option'] = alertsDF['Option'].str.strip()
    alertsDF['Strike'] = alertsDF['Strike'].str.strip()
    alertsDF.rename(columns={'Option': 'Symbol'}, inplace=True)

    #format the new option type & strike columns
    alertsDF['Option Type'] = alertsDF.apply(replaceText, axis = 1)
    alertsDF['Strike'] = alertsDF.apply(cleanStrike, axis=1)
    alertsDF['Strike'] = alertsDF['Strike'].str.replace(',', '')
    alertsDF['Strike'] = alertsDF['Strike'].astype('float')

    # Clean up Max Gain columns
    # split into individual $ and % columns
    # recast columns to ensure they are floats 
    # drop the old columns
    
    # if  $ index is lower than % index then $ first and % second 
    sample = alertsDF['High'][0]
    if sample.find('%') < sample.find('$'):
        alertsDF[['Max Gain %', 'Max Gain']] = alertsDF['High'].str.split(' ', expand=True)
    else:
        alertsDF[['Max Gain', 'Max Gain %']] = alertsDF['High'].str.split(' ', expand=True)
    
    alertsDF['Max Gain'] = alertsDF['Max Gain'].str.replace('$', '', regex=True)
    alertsDF['Max Gain'] = alertsDF['Max Gain'].str.replace(',', '', regex=True)
    alertsDF['Max Gain %'] = alertsDF['Max Gain %'].str.replace(')', '', regex=True)
    alertsDF['Max Gain %'] = alertsDF['Max Gain %'].str.replace('(', '', regex=True)
    alertsDF['Max Gain %'] = alertsDF['Max Gain %'].str.replace('%', '', regex=True)
    alertsDF['Max Gain %'] = alertsDF['Max Gain %'].str.replace(',', '', regex=True)
    alertsDF['Max Gain'] = alertsDF['Max Gain'].astype('float')
    alertsDF['Max Gain %'] = alertsDF['Max Gain %'].astype('float')

    # Clean up Max Loss columns
    # split into individual $ and % columns
    # recast columns to ensure they are floats
    sample = alertsDF['Low'][0]
    if sample.find('%') < sample.find('$'):
        alertsDF[['Max Loss %', 'Max Loss']] = alertsDF['Low'].str.split(' ', expand=True)
    else: 
        alertsDF[['Max Loss', 'Max Loss %']] = alertsDF['Low'].str.split(' ', expand=True)
    alertsDF['Max Loss'] = alertsDF['Max Loss'].str.replace('$', '', regex=True)
    alertsDF['Max Loss'] = alertsDF['Max Loss'].str.replace(',', '', regex=True)
    alertsDF['Max Loss %'] = alertsDF['Max Loss %'].str.replace(')', '', regex=True)
    alertsDF['Max Loss %'] = alertsDF['Max Loss %'].str.replace('(', '', regex=True)
    alertsDF['Max Loss %'] = alertsDF['Max Loss %'].str.replace('%', '', regex=True)
    #alertsDF['Max Gain %'] = alertsDF['Max Gain %'].str.replace(',', '', regex=True)
    alertsDF['Max Loss'] = alertsDF['Max Loss'].astype('float')
    alertsDF['Max Loss %'] = alertsDF['Max Loss %'].astype('float')

    # making sure no NaN's 
    alertsDF.fillna(0, inplace=True)

    # drop the old columns
    alertsDF = alertsDF.drop(columns=['High','Low', 'Time'], axis=1)

    #####
    # Adding compuited columns:
    # 'DTE' Date to Expiry of the option when the alert was first fired 
    alertsDF['DTE'] = (alertsDF['Expiry'] - alertsDF['Alert Date']).dt.days

    return alertsDF

#####
# Cleaninig up the slices to remove unnneeded columns i.e. strings 
# to prepare for KMeans analysis 
#####
def prepAlertsDataForKMeans(alertsDF, type='all'):

    dataframeOfFloats = pd.DataFrame()

    if type == 'all':
        dataframeOfFloats = alertsDF.drop( ['Symbol', 'Expiry', 'Max Loss', 'Option Type', 'Sector', 'Underlying', '% Diff', 'Ask', 'Alert Date', 'Tier', 'Max Gain' ], axis=1 )
    

    elif type == 'optionType':
        dataframeOfFloats = alertsDF.drop( ['Symbol', 'Expiry', 'Max Loss', 'Sector', 'Underlying', 'Ask', 'Alert Date', 'Tier', 'Max Gain' ], axis=1 )

    elif type == 'sector':
        dataframeOfFloats = alertsDF.drop(['Symbol', 'Expiry', 'Max Loss', 'Option Type', 'Underlying', '% Diff', 'Ask', 'Alert Date', 'Tier', 'Max Gain' ], axis=1 )

    return dataframeOfFloats

#####
# clustering - elbow method 
##
# Results: 
# All alerts -> 2 - 4
# over200 -> 2 - 5
#####
def elbowMethod(dataframeOfFloats):
    clusters = []

    for i in range(1, 11):
        km = KMeans(n_clusters=i).fit(dataframeOfFloats)
        clusters.append(km.inertia_)

    fig, ax = plt.subplots(figsize=(12,8))
    sns.lineplot(x=list(range(1,11)), y=clusters, ax=ax)
    ax.set_title('Searching for Elbow')
    ax.set_xlabel('Clusters')
    ax.set_ylabel('Intertia')

    plt.show()

#####
# Clustering & plotting
#####
def plotCluster(myDF, numClusters):
    km3 = KMeans(n_clusters = numClusters).fit(myDF)
    myDF['Labels'] = km3.labels_ 
    
    plt.figure(figsize=(12, 8))
    sns.scatterplot(myDF['Volume'], myDF['Max Gain %'], hue=myDF['Labels'],
    palette=sns.color_palette('hls', numClusters))
    plt.title('KMeans with 2 clusters')
    print(myDF.head())
    
    plt.show()

#####
# Returns a df with stats for the passed in alerts dataframe
# Alert dataset contains multiple or a single symbol
#####
def generalAlertStats(alertsDF1, sliceName = 'default', sheetName = 'default'):
    totalAlerts = alertsDF1['Expiry'].count()
    totalPostiveAlerts = alertsDF1.loc[alertsDF1['Max Gain %'] > 0]['Alert Date'].count()
    totalNegativeAlerts = alertsDF1.loc[alertsDF1['Max Gain %'] <= 0]['Alert Date'].count()

    # if the # of unique symbols > 1
    colName = ''
    if alertsDF1['Symbol'].nunique() == 1:
        colName = 'Symbol'
        colVal = alertsDF1.iloc[0]['Symbol']
    else:
        colName = 'Sheet Name'
        colVal = sheetName

    percentStatsData = {
        #'Symbol': [alertsDF['Symbol']],
        colName: [ colVal ],
        
        'Slice Name': [sliceName],
        
        'Period Start': [alertsDF1['Alert Date'].min().strftime('%Y-%m-%d')], 
        
        'Period End': [alertsDF1['Alert Date'].max().strftime('%Y-%m-%d')],
        
        'Total Alerts': [alertsDF1['Alert Date'].count()], 
        
        'Percent Positive': [totalPostiveAlerts / totalAlerts * 100],

        'Percent Above 100': [alertsDF1.loc[(alertsDF1['Max Gain %'] >= 100)]['Alert Date'].count() / totalAlerts * 100],
        
        'Percent Negative': [totalNegativeAlerts / totalAlerts * 100],
        
        'Gain % (Calls)': [alertsDF1.loc[alertsDF1['Option Type'] == 'Call']['Max Gain %'].mean()],
        
        'Gain % (Puts)': [alertsDF1.loc[alertsDF1['Option Type'] == 'Put']['Max Gain %'].mean()],
        
        '<0' : [alertsDF1.loc[(alertsDF1['Max Gain %'] <= 0)]['Alert Date'].count() / totalAlerts * 100],
        
        '0-20': [alertsDF1.loc[(alertsDF1['Max Gain %'] <= 20) & (alertsDF1['Max Gain %'] > 0)]['Alert Date'].count() / totalAlerts * 100],
        
        '21-50': [alertsDF1.loc[(alertsDF1['Max Gain %'] <= 50) & (alertsDF1['Max Gain %'] > 20)]['Alert Date'].count() / totalAlerts * 100],
        
        '51-100': [alertsDF1.loc[(alertsDF1['Max Gain %'] <= 100) & (alertsDF1['Max Gain %'] > 50)]['Alert Date'].count() / totalAlerts * 100],
        
        '101-200': [alertsDF1.loc[(alertsDF1['Max Gain %'] <= 200) & (alertsDF1['Max Gain %'] > 100)]['Alert Date'].count() / totalAlerts * 100],
        
        '201-500': [alertsDF1.loc[(alertsDF1['Max Gain %'] <= 500) & (alertsDF1['Max Gain %'] > 200)]['Alert Date'].count() / totalAlerts * 100],
        
        '500+': [alertsDF1.loc[(alertsDF1['Max Gain %'] > 500)]['Alert Date'].count() / totalAlerts * 100]

    }
    percentStats = pd.DataFrame(percentStatsData)
    percentStats = percentStats.round(2)
    return percentStats

#####
# Generates a dataframe that contains performance statistics 
# of different slices of the passed in alerts dataframe
##
# new DF = [ Symbol, Slice, stats pulled from generalAlertStat():::]
# if sheetName = default i.e. no sheetName is passed, the code assumes
# that the passed in alerts are for a single symbol
#####
def generateSliceStats(alertsDF, sheetName = 'default'):
    alertSlices = pd.DataFrame

    if alertsDF['Symbol'].nunique() == 1:
        alertSlices = generalAlertStats(alertsDF, 'baseline-%s'%(alertsDF['Symbol'][0]), sheetName=sheetName)
    else:
        alertSlices = generalAlertStats(alertsDF, 'baseline', sheetName=sheetName)
    
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Total $'] < 100000) & (alertsDF['OI'] < alertsDF['Volume']))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'$vol<100K;OI<Vol', sheetName=sheetName ))

    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Total $'] < 100000) & (alertsDF['IV'] < 40))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'$vol<100K;IV<40', sheetName=sheetName ))
    
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['DTE'] > 20))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'DTE>20', sheetName=sheetName ))
    
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['DTE'] < 20))]
    if not alertsDF_reduced.empty:
         alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'DTE<20', sheetName=sheetName ))

    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Tier'] == 'premium') & (alertsDF['DTE'] > 20))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'Premium; DTE>20', sheetName=sheetName ))
    
    # ask < 4;  volume < median;    % diff > 20
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Ask'] < 4) & (alertsDF['Volume'] > alertsDF['Volume'].median()) & (alertsDF['% Diff'] > 0.2) )]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'ask<4;vol>med;diff>20', sheetName=sheetName ))
    
    # Calls that are tagged as Bullish 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Call') & (alertsDF['Emojis'].str.contains("Bullish") ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'CallsBullish', sheetName=sheetName ))

    # Calls that are tagged as Bullish + Ask side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Call') & (alertsDF['Emojis'].str.contains("Bullish") & (alertsDF['Emojis'].str.contains("Ask Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'CallsBullishAskSide', sheetName=sheetName ))

    # Calls that are tagged as Bullish + Bid side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Call') & (alertsDF['Emojis'].str.contains("Bullish") & (alertsDF['Emojis'].str.contains("Bid Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'CallsBullishBidSide', sheetName=sheetName ))

    # Calls that are tagged as Bearish + Ask side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Call') & (alertsDF['Emojis'].str.contains("Bearish") & (alertsDF['Emojis'].str.contains("Ask Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'CallsBearishAskSide', sheetName=sheetName ))

    # Calls that are tagged as Bearish + Bid side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Call') & (alertsDF['Emojis'].str.contains("Bearish") & (alertsDF['Emojis'].str.contains("Bid Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'CallsBearishBidSide', sheetName=sheetName ))
    
    # Puts that are tagged as Bullish + Ask side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Put') & (alertsDF['Emojis'].str.contains("Bullish") & (alertsDF['Emojis'].str.contains("Ask Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'PutsBullishAskSide', sheetName=sheetName ))

    # Puts that are tagged as Bullish + Bid side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Put') & (alertsDF['Emojis'].str.contains("Bullish") & (alertsDF['Emojis'].str.contains("Bid Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'PutsBullishBidSide', sheetName=sheetName ))

    # Puts that are tagged as Bearish  
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Put') & (alertsDF['Emojis'].str.contains("Bearish") ) )]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'PutsBearish', sheetName=sheetName ))

    # Puts that are tagged as Bearish + Ask side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Put') & (alertsDF['Emojis'].str.contains("Bearish") & (alertsDF['Emojis'].str.contains("Ask Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'PutsBearishAskSide', sheetName=sheetName ))

    # Puts that are tagged as Bearish + Bid side 
    alertsDF_reduced = alertsDF.loc[ ( (alertsDF['Option Type'] == 'Put') & (alertsDF['Emojis'].str.contains("Bearish") & (alertsDF['Emojis'].str.contains("Bid Side")) ))]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'PutsBearishBidSide', sheetName=sheetName ))

    # Alert age < 5 days
    alertsDF_reduced = alertsDF.loc[ ((datetime.today() - alertsDF['Alert Date']).dt.days < 5)]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'alert<5', sheetName=sheetName ))
    
    # Alert age > 5 days
    alertsDF_reduced = alertsDF.loc[ ((datetime.today() - alertsDF['Alert Date']).dt.days > 5)]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'alert>5', sheetName=sheetName ))
    
    # Alert age > 10 days
    alertsDF_reduced = alertsDF.loc[ ((datetime.today() - alertsDF['Alert Date']).dt.days > 10)]
    if not alertsDF_reduced.empty:
        alertSlices = alertSlices.append(generalAlertStats(alertsDF_reduced,'alert>10', sheetName=sheetName ))

    # alertsDF['DTE'] = (alertsDF['Expiry'] - alertsDF['Alert Date']).dt.days

    return alertSlices.sort_values(by='Percent Above 100', ascending=False)

#####
# Takes in any Alerts dump and returns the alerts that are within the passed in sliceName
#####
def getSliceAlerts(alertsDF, sliceName):
    
    # IF THE SLICENAME CANNOT BE FOUND THEN DEFAULT BEHAVIOUR IS TO RETURN ALL ALERTS
    selectedSlice = alertsDF
    
    if sliceName == '$vol<100K;OI<Vol':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Total $'] < 100000) & (selectedSlice['OI'] < selectedSlice['Volume']))]
    
    elif sliceName == 'DTE>20':
        selectedSlice = selectedSlice.loc[ ( selectedSlice['DTE'] > 20)]

    elif sliceName == 'DTE<20':
        selectedSlice = selectedSlice.loc[ ( selectedSlice['DTE'] < 20)]
        
    elif sliceName == 'Premium; DTE>20':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Tier'] == 'premium') & (selectedSlice['DTE'] > 20))]
    
    elif sliceName == 'Premium; DTE<20':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Tier'] == 'premium') & (selectedSlice['DTE'] < 20))]
        
    elif sliceName == 'ask<4;vol>med;diff>20':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Ask'] < 4) & (selectedSlice['Volume'] > selectedSlice['Volume'].median()) & (selectedSlice['% Diff'] > 0.2) )]
    
    elif sliceName == '$vol<100K;IV<40':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Total $'] < 100000) & (selectedSlice['IV'] < 40))]
    
    elif sliceName == 'CallsBullish':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Call') & (selectedSlice['Emojis'].str.contains("Bullish")) )]

    elif sliceName == 'CallsBullishAskSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Call') & (selectedSlice['Emojis'].str.contains("Bullish") & (selectedSlice['Emojis'].str.contains("Ask Side")) ))]

    elif sliceName == 'CallsBullisBidSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Call') & (selectedSlice['Emojis'].str.contains("Bullish") & (selectedSlice['Emojis'].str.contains("Bid Side")) ))]
    
    elif sliceName == 'CallsBearishBidSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Call') & (selectedSlice['Emojis'].str.contains("Bearish") & (selectedSlice['Emojis'].str.contains("Bid Side")) ))]
    
    elif sliceName == 'CallsBearishAskSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Call') & (selectedSlice['Emojis'].str.contains("Bearish") & (selectedSlice['Emojis'].str.contains("Ask Side")) ))]
        
    elif sliceName == 'PutsBullishAskSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Put') & (selectedSlice['Emojis'].str.contains("Bullish") & (selectedSlice['Emojis'].str.contains("Ask Side")) ))]
        
    elif sliceName == 'PutsBullishBidSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Put') & (selectedSlice['Emojis'].str.contains("Bullish") & (selectedSlice['Emojis'].str.contains("Bid Side")) ))]

    elif sliceName == 'PutsBearish':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Put') & (selectedSlice['Emojis'].str.contains("Bearish") ))]

    elif sliceName == 'PutsBearishBidSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Put') & (selectedSlice['Emojis'].str.contains("Bearish") & (selectedSlice['Emojis'].str.contains("Bid Side")) ))]
    
    elif sliceName == 'PutsBearishAskSide':
        selectedSlice = selectedSlice.loc[ ( (selectedSlice['Option Type'] == 'Put') & (selectedSlice['Emojis'].str.contains("Bearish") & (selectedSlice['Emojis'].str.contains("Ask Side")) ))]
    
    elif sliceName == 'alert<5days':
        selectedSlice = selectedSlice.loc[ ((datetime.today() - selectedSlice['Alert Date']).dt.days < 5)]

    elif sliceName == 'alert>5days':
        selectedSlice = selectedSlice.loc[ ((datetime.today() - selectedSlice['Alert Date']).dt.days > 5)]

    elif sliceName == 'alert>10days':
        selectedSlice = selectedSlice.loc[ ((datetime.today() - selectedSlice['Alert Date']).dt.days > 10)]
    
    #selectedSlice.loc[:, 'Slice Name'] = sliceName
    return selectedSlice.sort_values(by='Max Gain %', ascending=False)

#####
# Look for the most frequently occuring symbols
#####
def findSymbolsWithHighFrequency(cleanAlertsDF):
    print('finding frequent Symbols!')
    # symbol, # alerts in last 3, 5, 8 days
    groupedAlerts = cleanAlertsDF.groupby(by='Symbol').count()
    print(groupedAlerts.head(10))

#####
# Plot Max Gain % as the following: 
# 1. Call/Put highs over time 
# 2. highs histogram (call/put combined)
#####
# TODO normalize the bins on the histogram so they are comparable between slices
def plotReturns(alertsDF_list, title="Calls vs. Puts"): 
    numRows = len(alertsDF_list) # Set the size of the figure
    xaxis_timeEnd = date.today()
    xaxis_timeStart = date(2021, 11, 22)

    with plt.style.context(("seaborn","ggplot")):
        fig = plt.figure(constrained_layout=True, figsize=(numRows * 3.2,10))
        specs = gridspec.GridSpec(ncols=2, nrows=numRows, figure=fig) ## Declaring 1xnumRows figure
        
        count = 0 
        for alertsDF in alertsDF_list:
            
            if alertsDF['Symbol'].nunique() == 1:
                xAxis_title = alertsDF['Symbol'][0] + alertsDF['Slice Name'].values[0]
            else:
                xAxis_title = alertsDF['Slice Name'].values[0]
            count += 1
            
            alertsDF.sort_values(by='Alert Date', inplace=True, ascending=False)
            calls = alertsDF.loc[alertsDF['Option Type'] == 'Call']
            puts = alertsDF.loc[alertsDF['Option Type'] == 'Put']
            
            # LINEGRAPH: plot returns over time, labeled by call and put
            x1 = fig.add_subplot(numRows, 2, count) # https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.subplot.html
            x1.plot( calls['Alert Date'], calls['Max Gain %'], color = 'g',label='Calls', marker='o')
            x1.plot( puts['Alert Date'], puts['Max Gain %'], color = 'r', label='Puts', marker='o')
            x1.set_title("%s - Plot"%(xAxis_title))
            x1.set_xlim( left = xaxis_timeStart, right = xaxis_timeEnd )

            # HISTOGRAM: returns
            count +=1 #bump up index in gridspec 
            numBins = 10 #math.ceil((alertsDF['Max Gain %'].max() - alertsDF['Max Gain %'].min())/3)
            x2 = fig.add_subplot(numRows, 2, count)
            #[-20, 0, 50, 100, 300, 500, 1500]
            x2.hist(alertsDF['Max Gain %'], color='tab:orange', bins=[-20, 0, 50, 100, 300, 500, 1500],  rwidth = 0.3, edgecolor='black')
            x2.set_title("%s - Histo"%(xAxis_title))

        plt.show()
        plt.close(fig)

#####
# Quick and dirty analysis of any list of alerts. Prints & Plots top slices
# input: clean dataframe of alerts
# output: slice stats, baseline alerts, baseline + top 4 slices 
#####
def quickAnalysis(alertsDF, sortby = 'Percent Above 100'):
    symbolAlerts = alertsDF 

    sliceStats = generateSliceStats(symbolAlerts) # get stats for the passed in alerts 
    sliceStats.sort_values(by=sortby, inplace=True, ascending=False) #sort by whats most important
    
    alertsInBestSlice = getSliceAlerts(symbolAlerts, sliceStats.iloc[0]['Slice Name']) # get teh alerts in the best slice for plotting later 
    baseline = getSliceAlerts(symbolAlerts, 'baseline') #always grab the baseline stats 
    baseline['Slice Name'] = 'baseline'

    listOfSlices = list()
    for n in range(4):
        alertsInSlice = getSliceAlerts(symbolAlerts, sliceStats.iloc[n]['Slice Name'])
        #alertsInSlice['Slice Name'] = sliceStats.iloc[n]['Slice Name'] #add slice name to the retrieved slice for later use
        alertsInSlice.reset_index(drop=True, inplace=True)
        alertsInSlice['Slice Name'] = sliceStats.iloc[n]['Slice Name']
        listOfSlices.append(alertsInSlice)
    
    print('')
    print('Printing Alerts in best slice...', sliceStats.iloc[0]['Slice Name'])
    print('')
    print(alertsInBestSlice.sort_values(by='Alert Date', ascending=False).head(30)[['Symbol', 'Strike', 'Option Type', 'Expiry', 'Ask', 'Max Gain %', 'Total $', 'IV', 'OI', 'Alert Date']])

    print('')
    print('Baseline Alerts...')
    print('')
    print(baseline.sort_values(by='Alert Date', ascending=False).head(30)[['Symbol', 'Strike', 'Option Type', 'Expiry', 'Ask', 'Max Gain %', 'Total $', 'IV', 'OI', 'Alert Date']])

    print('')
    print('Slices stats...')
    print(sliceStats)
    
    plotReturns([baseline, listOfSlices[0], listOfSlices[1], listOfSlices[2], listOfSlices[3]])

#####
# Objective: Visualize return characterisitcs of the top 5 slices for all available Symbol alerts 
###
# DONE - load all symbol alerts 
# compile slice stats for all 
# rank slice stats by % above 100 and weighted by # alerts 
# plot the return curves of the top 5 slices 
def compareAllSymbolAlerts(sortby = 'Percent Above 100'):
    alertsMap = getAlerts('all symbols') #returns dict of dataframes
    allSliceStats = pd.DataFrame()
    # TODO get BASELINE stats of all symbols 

    print(alertsMap.keys())
    for key in alertsMap: # generate slice stats for all symbols in the dict of dataframes 
        symbolAlerts = cleanAlertsData(alertsMap[key])
        symbolAlerts.loc[symbolAlerts['Max Gain %']>50.5]
        sliceStats = generateSliceStats(symbolAlerts)
        sliceStats = sliceStats.loc[sliceStats['Total Alerts'] > 10]
        allSliceStats = pd.concat([allSliceStats, sliceStats])
    
    # sort to get the top x slices 
    allSliceStats.sort_values(by=sortby, inplace=True, ascending=False)
    allSliceStats.reset_index(drop=True, inplace = True)
    
    # print the top 10 slices 
    print(allSliceStats.head(10))

    listOfSlices = list()
    # plot the top 5 slices
    for n in range(5):
        symbol = allSliceStats['Symbol'][n]
        alertsInSlice = getSliceAlerts( cleanAlertsData(alertsMap[symbol]), allSliceStats['Slice Name'][n] )
        alertsInSlice.reset_index(drop=True, inplace=True)
        alertsInSlice['Slice Name'] = allSliceStats['Slice Name'][n]
        listOfSlices.append(alertsInSlice)
    
    plotReturns( [listOfSlices[0], listOfSlices[1], listOfSlices[2], listOfSlices[3], listOfSlices[4] ])

    return listOfSlices

#####
# Objective: Spit out descriptive stats, and top slices for a large data dump 
# of general alerts from the UW Alerts feed
#####
def compareAllGeneralAlerts():
    print('analysis incomplete!')
    alertsDF = getAlerts('All Alerts')
    startDate = alertsDF['Alert Date'].astype(str).min()
    endDate = alertsDF['Alert Date'].astype(str).max()

    sliceStats = generateSliceStats(alertsDF, 'All Alerts')
    print('')
    print('Time Range = %s - %s (%s Days)'%(startDate, endDate, (alertsDF['Alert Date'].max() - alertsDF['Alert Date'].min()).days
    ))
    print('Total Alerts: %d'%(alertsDF['Alert Date'].count()))
    print('')
    sliceStats.reset_index(drop=True, inplace=True)
    print(sliceStats.head(10))

    listOfSlices = list()
    # plot the top 5 slices
    for n in range(5):
        sliceAlerts = getSliceAlerts(alertsDF, sliceStats['Slice Name'][n] )
     #   sliceAlerts.reset_index(drop=True, inplace=True)
        sliceAlerts['Slice Name'] = sliceStats['Slice Name'][n]
        listOfSlices.append(sliceAlerts)
    
    baseline = getSliceAlerts(alertsDF, 'baseline')
    baseline['Slice Name'] = 'baseline'
    
    #plotReturns( [baseline, listOfSlices[0], listOfSlices[1], listOfSlices[2], listOfSlices[3], listOfSlices[4] ])

    return sliceStats

#####
# Summarize basic stats by 'watchlist' 
# for each watchlist plt scatters comparing all cols 
#####
def analyzeMyHunt(alertsDF):
    # for each unique watchlist in the passed in alertsDF
    # #alerts, avg max gain, std deviation of max gain
    summary = alertsDF.groupby(by='Watchlist').agg({ 'Max Gain %' : ['count', 'mean', 'min', 'max', 'std'] })

    print(summary)

    #result = df.groupby('Type').agg({'top_speed(mph)': ['mean', 'min', 'max']})

#####
# Objective: Help find symbols with higher frequency of higher returns (where
# alerts are most recent).
# Prints statistical profile of general alerts based grouping by Symbols. 
#####
def frequencyAnalysis_genAlerts(alertsDF):
    print('im not ready yet')
    #groupedAlerts = cleanAlertsDF.groupby(by='Symbol').count()
    groupedAlerts = alertsDF.groupby(by='Symbol').agg({ 'Max Gain %' : ['count', 'max', 'min', 'mean', 'std'] })['Max Gain %'].sort_values(by='count', ascending=False)
    
    print('Alerts sorted by count, stats on Max Gain %: ')
    print(groupedAlerts.head(10))

#####
# Objective: Spit out a list of options that can be 
# consumed by Zorro to backtest the options signals  
# Use case 1: SPY - Chain, Alert Date, Ask
###
# input: clean list of alerts
# output: print to cmd, and print to csv 
#           Alert Date, Symbol, Type, Strike, Expiry, Ask
#####
def zorro_generateOptionSignal(alertsDF):
    print('zorro_generateOptionSignal - function not complete!!')

    alertsDF_simplified = alertsDF[['Alert Date', 'Symbol', 'Option Type', 'Strike', 'Expiry', 'Ask', 'Max Gain %']]
    printAlertsToFile(alertsDF_simplified, 'SPY-simple')
    #print(alertsDF_simplified)




###################################
##########                             COMMAND 
###################################

#listOfSlices = compareAllSymbolAlerts()
#sliceStatsForAllAlerts = compareAllGeneralAlerts()
#allAlerts = getAlerts('All Alerts')
#frequencyAnalysis_genAlerts(allAlerts)

#print('')
#print(listOfSlices[0])

myAlerts = getAlerts('My Hunt')
#quickAnalysis(myAlerts)
#print(myAlerts.columns)
#zorro_generateOptionSignal(myAlerts)

analyzeMyHunt(myAlerts)
#print(myAlerts.columns)

###################################
###################################

#####
# Plotting clusters
#####
#elbowMethod(X_floatsOnly)
#plotCluster(X_floatsOnly, 2)

#####
# Plotting 
#####
# Plots scatter of each column 
#sns.pairplot(baseline, aspect=1.5)
#sns.pairplot(alertsDF_reduced, hue='Tier', aspect=1.5 )
#plotReturns(30)

#plt.show()

#############################################   ##################
########################## UNUSED FUNCTIONS     ##################
#############################################   ##################
